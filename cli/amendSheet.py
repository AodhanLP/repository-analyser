import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from config import Config as c

# Check if the Excel file exists
try:
    workbook = openpyxl.load_workbook('LP Github repos.xlsx')
    # Do your amendments to the existing workbook
except FileNotFoundError:
    # If the file doesn't exist, create a new workbook
    workbook = Workbook()
    workbook.save('LP Github repos.xlsx')  # Save the new workbook

# Select the 'LP Github repos' sheet or create it if it doesn't exist
sheet_name = 'LP Github repos'
if sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
else:
    sheet = workbook.create_sheet(sheet_name)

# Define the headers in the first row of the sheet if it's a new sheet
if sheet.max_row == 1:
    headers = [
        'Repository Name',
        'Package Manager',
        'Dependency Management',
        'Semantic Release',
        'GHA',
        'Integration Suite (GHA)',
        'Concurrency Rule (GHA)',
        'Mend (GHA)'
    ]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

# Get the next available row
next_row = sheet.max_row + 1

# Iterate over the repositories
for repo_name in c.repos:
    # Get the data for the repository
    package_manager = c.get_package_manager(repo_name)
    dependency_management = c.get_dependency_management(repo_name)
    semantic_release = c.get_semantic_release(repo_name)
    gha = c.get_gha(repo_name)
    integration_suite = c.get_gha_integration(repo_name)
    concurrency_rule = c.get_gha_concurrency(repo_name)
    mend_gha = c.get_mend_gha(repo_name)

    # Print the values to debug
    print("Repository:", repo_name)
    print("Package Manager:", package_manager)
    print("Dependency Management:", dependency_management)
    print("Semantic Release:", semantic_release)
    print("GHA:", gha)
    print("Integration Suite (GHA):", integration_suite)
    print("Concurrency Rule (GHA):", concurrency_rule)
    print("Mend (GHA):", mend_gha)
    print("------------------------")

    # Write the data to the sheet
    sheet.cell(row=next_row, column=1, value=repo_name)
    sheet.cell(row=next_row, column=2, value=package_manager)
    sheet.cell(row=next_row, column=3, value=dependency_management)
    sheet.cell(row=next_row, column=4, value=semantic_release)
    sheet.cell(row=next_row, column=5, value=gha)
    sheet.cell(row=next_row, column=6, value=integration_suite)
    sheet.cell(row=next_row, column=7, value=concurrency_rule)
    sheet.cell(row=next_row, column=8, value=mend_gha)

    # Increment the row index
    next_row += 1

# Save the modified workbook
workbook.save('LP Github repos.xlsx')
