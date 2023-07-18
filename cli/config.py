import os
import json
import glob
import subprocess
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import pytz

class Config:
    
    #
    # Colored output
    #
    
    RED = '\033[91m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    CYAN = '\033[96m'
    RESET = '\033[0m'
    
    #
    # List of repositories to analyze
    #

    repos = [
        "adapt-ami",
        "adapt-builder",
        "adapt-builder-importer",
        "adapt-contrib-aboutUs",
        "adapt-contrib-articleBlockSlider",
        "adapt-contrib-assessmentSingleSubmit",
        "adapt-contrib-assistedLearning",
        "adapt-contrib-avatar",
        "adapt-contrib-branching",
        "adapt-contrib-brightcovePlayer",
        "adapt-contrib-carbonMenu",
        "adapt-contrib-certificate",
        "adapt-contrib-chatbot",
        "adapt-contrib-homeRedirect",
        "adapt-contrib-cinnamonMenu",
        "adapt-contrib-confidenceBasedMarking",
        "adapt-contrib-confirmExit",
        "adapt-contrib-coreMenu",
        "adapt-contrib-courseScore",
        "adapt-contrib-dyslexia-helper",
        "adapt-contrib-exhibit",
        "adapt-contrib-fiftyShades",
        "adapt-contrib-fillInTheBlanks",
        "adapt-contrib-flipcard",
        "adapt-contrib-fontAwesomeIcons",
        "adapt-contrib-fullScreenMenu",
        "adapt-contrib-googleAnalytics",
        "adapt-contrib-healthGauge",
        "adapt-contrib-hiddenHotspots",
        "lp-home",
        "lp-home-web",
        "adapt-contrib-homeRedirect",
        "adapt-contrib-hotspotMenu",
        "adapt-contrib-iconReplacement",
        "adapt-contrib-imageGallery",
        "adapt-contrib-interactiveVideo",
        "adapt-contrib-introOutro",
        "adapt-contrib-learningObjectives",
        "adapt-contrib-magazineMenu",
        "adapt-contrib-mathJax",
        "adapt-contrib-mcqResults",
        "adapt-contrib-mentor",
        "adapt-contrib-menuItemCompletion",
        "adapt-contrib-openTextInput",
        "adapt-contrib-openTextSummary",
        "adapt-contrib-pageFooterNavigation",
        "adapt-contrib-pageHeaderNavigation",
        "adapt-contrib-paginator",
        "adapt-contrib-references",
        "adapt-contrib-referencesList",
        "adapt-contrib-responsiveEmbeddedHelper",
        "adapt-contrib-responsiveIframe",
        "adapt-contrib-reveal",
        "adapt-contrib-salsaTheme",
        "adapt-contrib-scheduleMenu",
        "adapt-contrib-scriptInjector",
        "adapt-contrib-scrollPrompt",
        "adapt-contrib-spectrumMenu",
        "adapt-contrib-stacker",
        "adapt-contrib-suppressScore",
        "adapt-contrib-textToSpeech",
        "adapt-contrib-trueFalse",
        "adapt-contrib-xapiUtils",
        "adapt-contrib-yesNo",
        "adapt-test-automation",
        "alexa-trivia-quiz",
        "analytics-comparative-frontend",
        "anon-db",
        "app-switcher",
        "authoring-api",
        "automation",
        "automation-analytics",
        "automation-api",
        "automation-auth-cron",
        "automation-cache",
        "automation-cache-helper",
        "automation-data-cron",
        "automation-frontend",
        "automation-integration-processor",
        "automation-salesforce-webhook",
        "automation-sendgrid-webhook",
        "automation-statistics-cron",
        "automation-subscriptions-api",
        "automation-subscriptions-frontend",
        "automation-worker",
        "botkit",
        "ci-aws-iam-roles",
        "ci-ecr-poc",
        "ci-iam-poc",
        "CI-PoC-Jira-Environments",
        "ci-release-checks",
        "ci-workflow-test",
        "ci-workflows",
        "cicd-poc-redshift",
        "codeCoverage",
        "conversAI",
        "converse",
        "curatr",
        "curatr-bot",
        "curatr-example-confs",
        "curatr-rasa",
        "curatr-sid",
        "CuratrApp",
        "data-engineering-devops",
        "design-assets-cdn",
        "design-system",
        "design-tokens",
        "development-tooling",
        "docker-ui",
        "docs",
        "dolphin-test",
        "ecosystem",
        "EMR-ETL-Cron",
        "execute-on-all-mongo-instances",
        "gdpr",
        "gitdash",
        "HSTiles",
        "HT2-Labs-Website",
        "HT2_ML_Prototypes_Notebooks",
        "insights",
        "Internal-dw-Salesforce",
        "interview-scores",
        "interview-test",
        "ioc-css-qatests",
        "iq",
        "iq-clouds-code",
        "iq-dev-api",
        "iq-ops",
        "iq-performance-report",
        "iq-reporting",
        "iq-tools",
        "iq-ui-pro-dev",
        "jscommons",
        "kyle-sam-integration",
        "l5-core",
        "laravel-cascade-soft-deletes",
        "launchr",
        "learnio",
        "learnio-application",
        "learnio-automation",
        "learnio-backend-batch-job",
        "learnio-ddt-dev",
        "learnio-legacy",
        "learnio-ops",
        "learnio-pactools-dev",
        "learnio-sftp-scripts",
        "learnio-smoke-test",
        "learnio-sso",
        "learnio-titan-dev",
        "ll-coding-kata",
        "ll-connectors",
        "ll-enterprise",
        "ll-test-automation",
        "ll-v1-to-v2-migrator",
        "LL_Verb_Count_Converter",
        "lp-home",
        "lp-home-web",
        "lpqa-slackbot",
        "message-processor",
        "MongoSingleTenant",
        "monorepo-template",
        "morau",
        "mural",
        "MYSQLSource",
        "nightlies",
        "NLP-data-parser",
        "node-tnef",
        "nyse-clb-logs",
        "nyse-clb-logs",
        "nyse-client-central",
        "nyse-course-catalog",
        "nyse-data-reporting",
        "nyse-email-delivery-kit",
        "nyse-email-delivery-kit-realbiz",
        "nyse-learn-io-pulse",
        "nyse-learn-io-site",
        "nyse-moodle",
        "nyse-pulse-check-leaderboard",
        "nyse-roi-calculator",
        "nyse-shareholder-activism-playbook",
        "nyse-toutils",
        "nyse-web-svn",
        "OpenID-Connect-PHP",
        "php-tech-test",
        "plop-generators",
        "poet",
        "poet-dev",
        "poet-ops",
        "poet-test-automation",
        "random-scripts",
        "react-native-http-bridge",
        "release-notes-helper",
        "releasing",
        "renovate-config",
        "rf-base-project",
        "RoyalCollegeOfNursing_Analytics",
        "rp_framework",
        "SAAS_DataLake",
        "SalesforceETL",
        "scholar",
        "scholar-course-select-dev",
        "scholar-docker",
        "scholar-game-flash",
        "scholar-game-library",
        "scholar-gdp",
        "scholar-gdp-game-test",
        "scholar-gdp-legacy",
        "scholar-modular-layouts",
        "scholar-mounted-disk",
        "scholar-release-notes",
        "scholar-script-parser",
        "scholar-test-automation",
        "scholar-tools-packages",
        "scholar-tools-pulse-check",
        "scholar-tools-python",
        "scholar-web-api",
        "scholar-wrapper-api",
        "semantic-analysis-node-test",
        "semantic-release",
        "semantic-release-demo",
        "serverless-cop-starter",
        "simplesamlphp",
        "sisense-config",
        "sisense-demo-2",
        "SisenseAPICalls",
        "sl-customer-auth-ui",
        "sl-financial-health-check",
        "sl-retirement-calculator",
        "sl-retirement-calculator-legacy",
        "sl-retirement-calculator-testing",
        "sl-retirement-pathfinder",
        "sl-robo-tools",
        "sl-total-rewards-statement",
        "sonarqube",
        "stream-analytics-test-automation",
        "stream-app",
        "stream-docker-dev",
        "stream-ui-qatests",
        "streamhome-test-automation",
        "teamsales",
        "test",
        "testProject",
        "TinCanPHP",
        "tol-hyrule",
        "translations-web-app",
        "typescript-project",
        "udd",
        "unity-xapi",
        "upptime-status",
        "Valvoline_Analytics",
        "vcs-migration-tooling",
        "venue",
        "waves-test-automation",
        "weasal-python",
        "whitesource-configs",
        "xAI",
        "xAI-API",
        "xAI-Console-UI",
        "xAI-Inference-server",
        "xAI-Labelling-UI",
        "xapi-js-client",
        "yum-api-qatests"
    ]

    #
    # Functions
    #
    
    def load_or_create_workbook(file_path):
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = Workbook()
        return workbook

    def select_or_create_sheet(workbook, sheet_name):
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
        return sheet

    def write_headers(sheet, headers):
        if sheet.max_row == 1:
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)

    # Clone the repository and change directory
    def clone_repo(repo_name):
        try:
            os.system(f"git clone https://github.com/HT2-Labs/{repo_name}.git")
            os.chdir(repo_name)
        except:
            print(f'Failed to clone the repository {repo_name}.')
            print('Exiting program.')
            exit()

    # Change directory back to root and delete the cloned repository
    def return_to_root(repo_name):
        os.chdir("..")
        os.system(f"rm -rf {repo_name}")

    # Check the repository for 'package-lock.json' or 'yarn.lock'
    def get_package_manager(repo_name):
        try:
            # Check if 'package-lock.json' and 'yarn.lock' exist in the repository
            packageLockJsonExists = os.path.isfile('package-lock.json')
            yarnLockExists = os.path.isfile('yarn.lock')

            if packageLockJsonExists and yarnLockExists:
                return 'Yes (NPM and Yarn)' 
            elif packageLockJsonExists:
                return 'Yes (NPM)'
            elif yarnLockExists:
                return 'Yes (Yarn)'
            else:
                return 'No'
        except:
            print(f'Failed to check for a package manager for {repo_name}.')
            print('Exiting program.')
            exit()

    # Check the repository for Semantic Release
    def get_semantic_release(repo_name):
        try:
            packageJsonExists = os.path.isfile('package.json')

            if packageJsonExists:
                with open('package.json', 'r') as f:
                    data = json.load(f)

                semanticRelease = 'devDependencies' in data and 'semantic-release' in data['devDependencies']
            else:
                semanticRelease = False

            # If semanticRelease is True, return 'Yes'. Otherwise, return 'No'
            if semanticRelease:
                return 'Yes'
            else:
                return 'No'
        except:
            print(f'Failed to check for Semantic Release for {repo_name}.')
            print('Exiting program.')
            exit()
            
    # Check the repository for Integration Suite (GHA)
    def get_gha_integration(repo_name):
        try:
            ymlFiles = glob.glob('.github/workflows/*.yml')
            yamlFiles = glob.glob('.github/workflows/*.yaml')
            allFiles = ymlFiles + yamlFiles
            keywords = ["integration suite", "integration"]
            for file in allFiles:
                with open(file, 'r') as f:
                    if any(keyword in line for line in f for keyword in keywords):
                        return "Yes"
            return "No"
        except:
            print(f'Failed to check for Integration Suite (GHA) for {repo_name}.')
            print('Exiting program.')
            exit()

    # Check the repository for Concurrency Rule (GHA)
    def get_gha_concurrency(repo_name):
        try:
            ymlFiles = glob.glob('.github/workflows/*.yml')
            yamlFiles = glob.glob('.github/workflows/*.yaml')
            allFiles = ymlFiles + yamlFiles
            for file in allFiles:
                with open(file, 'r') as f:
                    if 'concurrency' in f.read():
                        return "Yes"
            return "No"
        except:
            print(f'Failed to check for Concurrency Rule (GHA) for {repo_name}.')
            print('Exiting program.')
            exit()

    # Check the repository for Mend (GHA)
    def get_mend_gha(repo_name):
        try:
            ymlFiles = glob.glob('.github/workflows/*.yml')
            yamlFiles = glob.glob('.github/workflows/*.yaml')
            allFiles = ymlFiles + yamlFiles
            for file in allFiles:
                with open(file, 'r') as f:
                    if 'mend' in f.read():
                        return "Yes"
            return "No"
        except:
            print(f'Failed to check for Mend (GHA) for {repo_name}.')
            print('Exiting program.')
            exit()

    # Check the repository for Dependency Management
    def get_dependency_management(repo_name):
        try:
            process = subprocess.run(["gh", "pr", "list"], capture_output=True, text=True)
            output = process.stdout
            if "renovate" and "dependabot" in output :
                return "Renovate and Dependabot"    
            elif "dependabot" in output:
                return "Dependabot"
            elif "renovate" in output:
                return "Renovate"
            return "No"
        except:
            print(f'Failed to check for Dependency Management for {repo_name}.')
            print('Exiting program.')
            exit()

    # Check the repository for GitHub Actions files
    def get_gha(repo_name):
        try:
            ymlFiles = glob.glob('.github/workflows/*.yml')
            yamlFiles = glob.glob('.github/workflows/*.yaml')

            files = ymlFiles + yamlFiles

            one_year_ago = datetime.datetime.now(pytz.utc) - datetime.timedelta(days=365)

            for file in files:
                result = subprocess.run(['git', 'log', '-1', '--format=%cd', file], capture_output=True, text=True)
                file_date = datetime.datetime.strptime(result.stdout.strip(), '%a %b %d %H:%M:%S %Y %z').replace(tzinfo=pytz.utc)

                if file_date > one_year_ago:
                    return 'Yes'
            return 'No'  # If no recent files are found

        except:
            print(f'Failed to check for GitHub Actions files for {repo_name}.')
            print('Exiting program.')
            exit()
    
    def update_or_add_repo(sheet, repo, repo_exists, package_manager, dependency_management, semantic_release, gha, integration_suite, concurrency_rule, mend_gha):
        if repo_exists:
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=1).value == repo:
                    sheet.cell(row=row, column=2, value=package_manager)
                    sheet.cell(row=row, column=3, value=dependency_management)
                    sheet.cell(row=row, column=4, value=semantic_release)
                    sheet.cell(row=row, column=5, value=gha)
                    sheet.cell(row=row, column=6, value=integration_suite)
                    sheet.cell(row=row, column=7, value=concurrency_rule)
                    sheet.cell(row=row, column=8, value=mend_gha)
                    break
        else:
            next_row = sheet.max_row + 1
            sheet.cell(row=next_row, column=1, value=repo)
            sheet.cell(row=next_row, column=2, value=package_manager)
            sheet.cell(row=next_row, column=3, value=dependency_management)
            sheet.cell(row=next_row, column=4, value=semantic_release)
            sheet.cell(row=next_row, column=5, value=gha)
            sheet.cell(row=next_row, column=6, value=integration_suite)
            sheet.cell(row=next_row, column=7, value=concurrency_rule)
            sheet.cell(row=next_row, column=8, value=mend_gha)

            
    def console_output(repo, package_manager, semantic_release, gha, dependency_management, integration_suite, concurrency_rule, mend_gha):
        print(f'{Config.CYAN}--------------------------------{Config.RESET}')
        print(f'Repository: {Config.YELLOW}{repo}{Config.RESET}')
        print(f'Package Manager: {Config.RED if package_manager == "No" else Config.GREEN}{package_manager}{Config.RESET}')
        print(f'Semantic Release: {Config.GREEN if semantic_release == "Yes" else Config.RED}{semantic_release}{Config.RESET}')
        print(f'GitHub Actions: {Config.GREEN if gha == "Yes" else Config.RED}{gha}{Config.RESET}')
        print(f'Dependency Management: {Config.RED if dependency_management == "No" else Config.GREEN}{dependency_management}{Config.RESET}')
        print(f'Integration Suite (GHA): {Config.GREEN if integration_suite == "Yes" else Config.RED}{integration_suite}{Config.RESET}')
        print(f'Concurrency Rule (GHA): {Config.GREEN if concurrency_rule == "Yes" else Config.RED}{concurrency_rule}{Config.RESET}')
        print(f'Mend (GHA): {Config.GREEN if mend_gha == "Yes" else Config.RED}{mend_gha}{Config.RESET}')
        print(f'{Config.CYAN}--------------------------------{Config.RESET}')