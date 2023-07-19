import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from config import Config as c

# Check if the Excel file exists
workbook = c.load_or_create_workbook('LP Github repos.xlsx')

# Select the 'LP Github repos' sheet or create it if it doesn't exist
sheet_name = 'LP Github repos'
sheet = c.select_or_create_sheet(workbook, sheet_name)

# Define the headers in the first row of the sheet if it's a new sheet
if sheet.max_row == 1:
    headers = [
        'Repository Name',
        'Package Manager',
        'Dependency Management',
        'Semantic Release',
        'GHA',
        'Integration Suite (GHA)',
        'Concurrency Rule (GHA)',
        'Mend (GHA)'
    ]

    # Write the headers to the sheet
    c.write_headers(sheet, headers)

# Analyze each repository
for repo in c.repos:
    try:
        # Analyze the repository
        c.clone_repo(repo)
        package_manager = c.get_package_manager(repo)
        dependency_management = c.get_dependency_management(repo)
        semantic_release = c.get_semantic_release(repo)
        gha = c.get_gha(repo)
        integration_suite = c.get_gha_integration(repo)
        concurrency_rule = c.get_gha_concurrency(repo)
        mend_gha = c.get_mend_gha(repo)
        c.return_to_root(repo)

        # Check if the repository already exists in the sheet
        repo_exists = False
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == repo:
                repo_exists = True
                break

        # Update the data if the repository already exists, or add a new row otherwise
        c.update_or_add_repo(sheet, repo, repo_exists, package_manager, dependency_management, semantic_release, gha, integration_suite, concurrency_rule, mend_gha)

        # Save the modified workbook
        workbook.save('LP Github repos.xlsx')

        # Print the repository information
        c.console_output(repo, package_manager, semantic_release, gha, dependency_management, integration_suite,
                         concurrency_rule, mend_gha)

    except:
        print(f'{c.RED}Failed to analyze {repo}.{c.RESET}')
        print('Exiting program.')
        exit()